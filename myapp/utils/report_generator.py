"""
Sales Report Generator — Excel Only
====================================
Generates comprehensive Excel report with:
- Item Summary (one row per item with all sales combined)
- Market Share calculation
- Professional formatting
- Urdu/English support
"""

import os
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


async def generate_report_files(report_id, sales):
    """Generate Excel report with all sales data and item summaries."""
    
    unique_items = {s.item_name for s in sales}
    if len(unique_items) < 5:
        return {"error": True,
                "message": "رپورٹ جنریٹ کرنے کے لیے کم از کم 5 مختلف اشیاء کی فروخت ضروری ہے۔"}

    folder = f"reports_storage/report_{report_id}"
    os.makedirs(folder, exist_ok=True)

    # Prepare sales data
    sales_list = []
    for s in sales:
        up = float(s.unit_price) if s.unit_price else 0
        sales_list.append({
            "sale_id": s.sale_id,
            "item_name": s.item_name,
            "quantity_sold": s.quantity_sold,
            "unit_price": up,
            "item_unit": s.item_unit or "-",
            "sale_date": s.sale_date,
            "total_amount": s.quantity_sold * up,
            "customer_name": getattr(s, "customer_name", "عام کسٹمر"),
        })

    # Calculate totals
    total_qty = sum(s["quantity_sold"] for s in sales_list)
    total_rev = sum(s["total_amount"] for s in sales_list)
    total_trans = len(sales_list)
    total_items = len(unique_items)
    
    # Calculate average per transaction
    avg_per_transaction = total_rev / total_trans if total_trans > 0 else 0

    # Group by item for summary (ONE ROW PER ITEM)
    item_sales = defaultdict(lambda: {
        "quantity": 0, 
        "revenue": 0, 
        "count": 0,
        "unit": "-",
        "min_price": float('inf'),
        "max_price": 0,
        "prices": []
    })
    
    for s in sales_list:
        item_sales[s["item_name"]]["quantity"] += s["quantity_sold"]
        item_sales[s["item_name"]]["revenue"] += s["total_amount"]
        item_sales[s["item_name"]]["count"] += 1
        item_sales[s["item_name"]]["unit"] = s["item_unit"]
        item_sales[s["item_name"]]["prices"].append(s["unit_price"])
        if s["unit_price"] < item_sales[s["item_name"]]["min_price"]:
            item_sales[s["item_name"]]["min_price"] = s["unit_price"]
        if s["unit_price"] > item_sales[s["item_name"]]["max_price"]:
            item_sales[s["item_name"]]["max_price"] = s["unit_price"]
    
    # Calculate average price for each item
    for item in item_sales:
        if item_sales[item]["quantity"] > 0:
            item_sales[item]["avg_price"] = item_sales[item]["revenue"] / item_sales[item]["quantity"]
        else:
            item_sales[item]["avg_price"] = 0
    
    # Calculate market share
    total_quantity_all = sum(data["quantity"] for data in item_sales.values())
    
    # Sort items by quantity for top items
    top_5 = sorted(item_sales.items(), key=lambda x: x[1]["quantity"], reverse=True)[:5]

    # Create Excel file
    excel_path = os.path.join(folder, "sales_report.xlsx")
    wb = Workbook()
    
    # Style definitions
    header_font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
    header_fill = PatternFill(start_color="2C7DA0", end_color="2C7DA0", fill_type="solid")
    subheader_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment_center = Alignment(horizontal="center", vertical="center")
    cell_alignment_right = Alignment(horizontal="right", vertical="center")
    cell_alignment_left = Alignment(horizontal="left", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # =============================================================
    # SHEET 1: اشیاء کا خلاصہ (Items Summary - One row per item)
    # =============================================================
    ws1 = wb.active
    ws1.title = "اشیاء کا خلاصہ"
    
    # Add Report Header
    ws1.merge_cells('A1:J1')
    header_cell = ws1.cell(row=1, column=1)
    header_cell.value = f"📊 فروخت رپورٹ - {datetime.now().strftime('%Y-%m-%d')}"
    header_cell.font = Font(bold=True, size=16, color="1a5a7a", name="Arial")
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws1.merge_cells('A2:J2')
    subheader_cell = ws1.cell(row=2, column=1)
    subheader_cell.value = "تمام اشیاء کا تفصیلی جائزہ (ہر آئٹم کی تمام فروخت ایک قطار میں)"
    subheader_cell.font = Font(size=10, italic=True, color="666666", name="Arial")
    subheader_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Item Summary Headers
    summary_headers = [
        "درجہ", "آئٹم کا نام", "یونٹ", "کل فروخت مقدار", 
        "کل آمدنی (Rs.)", "فروخت کی تعداد", "اوسط قیمت (Rs.)",
        "مارکیٹ شیئر (%)", "کم ترین قیمت", "اعلیٰ ترین قیمت"
    ]
    
    for col, header in enumerate(summary_headers, 1):
        cell = ws1.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add item summary data
    for rank, (item_name, data) in enumerate(sorted(item_sales.items(), key=lambda x: x[1]["quantity"], reverse=True), 1):
        market_share = (data["quantity"] / total_quantity_all * 100) if total_quantity_all > 0 else 0
        row_num = 4 + rank
        
        ws1.cell(row=row_num, column=1, value=rank)
        ws1.cell(row=row_num, column=2, value=item_name)
        ws1.cell(row=row_num, column=3, value=data["unit"])
        ws1.cell(row=row_num, column=4, value=int(data["quantity"]))
        ws1.cell(row=row_num, column=5, value=round(data["revenue"], 2))
        ws1.cell(row=row_num, column=6, value=data["count"])
        ws1.cell(row=row_num, column=7, value=round(data["avg_price"], 2))
        ws1.cell(row=row_num, column=8, value=round(market_share, 2))
        ws1.cell(row=row_num, column=9, value=round(data["min_price"], 2) if data["min_price"] != float('inf') else 0)
        ws1.cell(row=row_num, column=10, value=round(data["max_price"], 2))
        
        for col in range(1, 11):
            cell = ws1.cell(row=row_num, column=col)
            cell.border = border
            if col in [4, 5, 7, 8, 9, 10]:
                cell.alignment = cell_alignment_right
            elif col == 2:
                cell.alignment = cell_alignment_left
            else:
                cell.alignment = cell_alignment_center
            
            if col == 8 and market_share > 20:
                cell.font = Font(bold=True, color="10b981")
            elif col == 8 and market_share > 10:
                cell.font = Font(bold=True, color="f59e0b")
    
    # Add totals row
    total_row = 4 + len(item_sales) + 1
    ws1.merge_cells(f'A{total_row}:C{total_row}')
    total_label = ws1.cell(row=total_row, column=1)
    total_label.value = "کل"
    total_label.font = Font(bold=True, size=12)
    total_label.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    total_label.alignment = Alignment(horizontal="center")
    
    ws1.cell(row=total_row, column=4, value=int(total_qty))
    ws1.cell(row=total_row, column=5, value=round(total_rev, 2))
    ws1.cell(row=total_row, column=6, value=total_trans)
    
    for col in [4, 5, 6]:
        cell = ws1.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        cell.border = border
        cell.alignment = cell_alignment_right
    
    for col in ws1.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 3, 25)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # =============================================================
    # SHEET 2: تمام فروخت کے ریکارڈز (All Sales Records)
    # =============================================================
    ws2 = wb.create_sheet("تمام فروخت کے ریکارڈز")
    
    ws2.merge_cells('A1:H1')
    header_cell2 = ws2.cell(row=1, column=1)
    header_cell2.value = "📋 تمام فروخت کے تفصیلی ریکارڈز"
    header_cell2.font = Font(bold=True, size=16, color="1a5a7a")
    header_cell2.alignment = Alignment(horizontal="center")
    
    ws2.merge_cells('A2:H2')
    subheader_cell2 = ws2.cell(row=2, column=1)
    subheader_cell2.value = "ہر فروخت کا الگ ریکارڈ (تاریخ، مقدار، قیمت کے ساتھ)"
    subheader_cell2.font = Font(size=10, italic=True, color="666666")
    subheader_cell2.alignment = Alignment(horizontal="center")
    
    sales_headers = ["شناختی نمبر", "آئٹم کا نام", "مقدار", "فی یونٹ قیمت (Rs.)", "یونٹ", "تاریخ", "کل رقم (Rs.)", "کسٹمر کا نام"]
    
    for col, header in enumerate(sales_headers, 1):
        cell = ws2.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for idx, s in enumerate(sales_list, 1):
        row_num = 4 + idx
        ws2.cell(row=row_num, column=1, value=s["sale_id"])
        ws2.cell(row=row_num, column=2, value=s["item_name"])
        ws2.cell(row=row_num, column=3, value=int(s["quantity_sold"]))
        ws2.cell(row=row_num, column=4, value=round(s["unit_price"], 2))
        ws2.cell(row=row_num, column=5, value=s["item_unit"])
        ws2.cell(row=row_num, column=6, value=s["sale_date"].strftime("%Y-%m-%d") if hasattr(s["sale_date"], "strftime") else str(s["sale_date"]))
        ws2.cell(row=row_num, column=7, value=round(s["total_amount"], 2))
        ws2.cell(row=row_num, column=8, value=s["customer_name"])
        
        for col in range(1, 9):
            cell = ws2.cell(row=row_num, column=col)
            cell.border = border
            if col in [3, 4, 7]:
                cell.alignment = cell_alignment_right
            elif col == 2:
                cell.alignment = cell_alignment_left
            else:
                cell.alignment = cell_alignment_center
    
    for col in ws2.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 3, 30)
        ws2.column_dimensions[column_letter].width = adjusted_width
    
    # =============================================================
    # SHEET 3: کارکردگی کے اشارے (KPI Dashboard)
    # =============================================================
    ws3 = wb.create_sheet("کارکردگی کے اشارے")
    
    ws3.merge_cells('A1:C1')
    header_cell3 = ws3.cell(row=1, column=1)
    header_cell3.value = "🎯 کارکردگی کے کلیدی اشارے (KPIs)"
    header_cell3.font = Font(bold=True, size=16, color="1a5a7a")
    header_cell3.alignment = Alignment(horizontal="center")
    
    ws3.merge_cells('A2:C2')
    subheader_cell3 = ws3.cell(row=2, column=1)
    subheader_cell3.value = "کاروبار کی مجموعی کارکردگی کے اہم میٹرکس"
    subheader_cell3.font = Font(size=10, italic=True, color="666666")
    subheader_cell3.alignment = Alignment(horizontal="center")
    
    kpi_data = [
        ["کی پی آئی", "قدر", "وضاحت"],
        ["کل فروخت مقدار", f"{total_qty:,}", "تمام اشیاء کی کل فروخت شدہ مقدار"],
        ["کل آمدنی", f"Rs. {total_rev:,.2f}", "کل فروخت سے حاصل کردہ رقم"],
        ["کل لین دین", f"{total_trans:,}", "کل فروخت کے ریکارڈز کی تعداد"],
        ["منفرد اشیاء", f"{total_items:,}", "مختلف اشیاء کی تعداد"],
        ["اوسط فروخت فی لین دین", f"Rs. {avg_per_transaction:,.2f}", "فی لین دین اوسط رقم"],
        ["کل مختلف کسٹمرز", f"{len(set(s.get('customer_name', '') for s in sales_list if s.get('customer_name'))):,}", "منفرد کسٹمرز کی تعداد"],
        ["سب سے زیادہ فروخت شدہ آئٹم", top_5[0][0] if top_5 else "-", f"{top_5[0][1]['quantity']:,} یونٹس" if top_5 else "-"],
        ["کل مارکیٹ ویلیو", f"Rs. {total_rev:,.2f}", "تمام فروخت کی کل مالیت"],
    ]
    
    for row_idx, row_data in enumerate(kpi_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            
            if row_idx == 4:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            else:
                if col_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = subheader_fill
                    cell.alignment = cell_alignment_left
                elif col_idx == 2:
                    cell.font = Font(bold=True, color="2C7DA0")
                    cell.alignment = cell_alignment_right
                else:
                    cell.alignment = cell_alignment_left
    
    for col in ws3.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 3, 40)
        ws3.column_dimensions[column_letter].width = adjusted_width
    
    # =============================================================
    # SHEET 4: بہترین 5 اشیاء (Top 5 Items)
    # =============================================================
    ws4 = wb.create_sheet("بہترین 5 اشیاء")
    
    ws4.merge_cells('A1:F1')
    header_cell4 = ws4.cell(row=1, column=1)
    header_cell4.value = "🏆 سب سے زیادہ فروخت ہونے والی بہترین 5 اشیاء"
    header_cell4.font = Font(bold=True, size=16, color="1a5a7a")
    header_cell4.alignment = Alignment(horizontal="center")
    
    ws4.merge_cells('A2:F2')
    subheader_cell4 = ws4.cell(row=2, column=1)
    subheader_cell4.value = "مارکیٹ شیئر اور کارکردگی کے ساتھ تفصیلی تجزیہ"
    subheader_cell4.font = Font(size=10, italic=True, color="666666")
    subheader_cell4.alignment = Alignment(horizontal="center")
    
    top_headers = ["درجہ", "آئٹم کا نام", "فروخت مقدار", "کل آمدنی (Rs.)", "مارکیٹ شیئر (%)", "فروخت کی تعداد"]
    
    for col, header in enumerate(top_headers, 1):
        cell = ws4.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for rank, (item_name, data) in enumerate(top_5, 1):
        market_share = (data["quantity"] / total_quantity_all * 100) if total_quantity_all > 0 else 0
        row_num = 4 + rank
        
        ws4.cell(row=row_num, column=1, value=rank)
        ws4.cell(row=row_num, column=2, value=item_name)
        ws4.cell(row=row_num, column=3, value=int(data["quantity"]))
        ws4.cell(row=row_num, column=4, value=round(data["revenue"], 2))
        ws4.cell(row=row_num, column=5, value=round(market_share, 2))
        ws4.cell(row=row_num, column=6, value=data["count"])
        
        for col in range(1, 7):
            cell = ws4.cell(row=row_num, column=col)
            cell.border = border
            if col in [3, 4, 5]:
                cell.alignment = cell_alignment_right
            elif col == 2:
                cell.alignment = cell_alignment_left
            else:
                cell.alignment = cell_alignment_center
            
            if col == 5:
                if market_share > 20:
                    cell.font = Font(bold=True, color="10b981")
                elif market_share > 10:
                    cell.font = Font(bold=True, color="f59e0b")
    
    for col in ws4.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 3, 25)
        ws4.column_dimensions[column_letter].width = adjusted_width
    
    # =============================================================
    # SHEET 5: روزانہ فروخت کا رجحان (Daily Sales Trend - IMPORTANT SHEET)
    # =============================================================
    ws5 = wb.create_sheet("روزانہ فروخت کا رجحان")
    
    ws5.merge_cells('A1:E1')
    header_cell5 = ws5.cell(row=1, column=1)
    header_cell5.value = "📈 روزانہ فروخت کا رجحان (Daily Sales Trend Analysis)"
    header_cell5.font = Font(bold=True, size=16, color="1a5a7a")
    header_cell5.alignment = Alignment(horizontal="center")
    
    ws5.merge_cells('A2:E2')
    subheader_cell5 = ws5.cell(row=2, column=1)
    subheader_cell5.value = "یہ شیٹ ظاہر کرتی ہے کہ وقت کے ساتھ فروخت کیسے بدل رہی ہے - کاروبار میں اضافہ یا کمی کا تجزیہ"
    subheader_cell5.font = Font(size=10, italic=True, color="666666")
    subheader_cell5.alignment = Alignment(horizontal="center")
    
    ws5.merge_cells('A3:E3')
    info_cell5 = ws5.cell(row=3, column=1)
    info_cell5.value = "💡 رجحان کا مطلب: اگر مقدار بڑھ رہی ہے تو کاروبار ترقی کر رہا ہے، اگر کم ہو رہی ہے تو وجہ معلوم کریں"
    info_cell5.font = Font(size=9, color="2C7DA0", italic=True)
    info_cell5.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    info_cell5.alignment = Alignment(horizontal="center")
    
    # Group sales by date
    daily_sales = defaultdict(lambda: {"quantity": 0, "revenue": 0, "count": 0})
    for s in sales_list:
        date_str = s["sale_date"].strftime("%Y-%m-%d") if hasattr(s["sale_date"], "strftime") else str(s["sale_date"])
        daily_sales[date_str]["quantity"] += s["quantity_sold"]
        daily_sales[date_str]["revenue"] += s["total_amount"]
        daily_sales[date_str]["count"] += 1
    
    # Sort by date
    sorted_dates = sorted(daily_sales.keys())
    
    # Trend headers with better descriptions
    trend_headers = [
        "تاریخ\n(Date)", 
        "کل فروخت مقدار\n(Total Quantity)", 
        "کل آمدنی (Rs.)\n(Total Revenue)", 
        "لین دین کی تعداد\n(Transaction Count)",
        "اوسط فی لین دین (Rs.)\n(Avg per Transaction)"
    ]
    
    for col, header in enumerate(trend_headers, 1):
        cell = ws5.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add daily data
    for idx, date in enumerate(sorted_dates, 1):
        data = daily_sales[date]
        avg_per_day_transaction = data["revenue"] / data["count"] if data["count"] > 0 else 0
        row_num = 5 + idx
        
        ws5.cell(row=row_num, column=1, value=date)
        ws5.cell(row=row_num, column=2, value=int(data["quantity"]))
        ws5.cell(row=row_num, column=3, value=round(data["revenue"], 2))
        ws5.cell(row=row_num, column=4, value=data["count"])
        ws5.cell(row=row_num, column=5, value=round(avg_per_day_transaction, 2))
        
        for col in range(1, 6):
            cell = ws5.cell(row=row_num, column=col)
            cell.border = border
            if col in [2, 3, 4, 5]:
                cell.alignment = cell_alignment_right
            else:
                cell.alignment = cell_alignment_center
            
            # Highlight high sales days
            if col == 3 and data["revenue"] > (total_rev / len(sorted_dates)) if sorted_dates else 0:
                cell.font = Font(bold=True, color="10b981")
    
    # Add trend analysis summary
    if sorted_dates:
        analysis_row = 5 + len(sorted_dates) + 2
        
        ws5.merge_cells(f'A{analysis_row}:E{analysis_row}')
        analysis_title = ws5.cell(row=analysis_row, column=1)
        analysis_title.value = "📊 رجحان کا تجزیہ (Trend Analysis)"
        analysis_title.font = Font(bold=True, size=12, color="1a5a7a")
        analysis_title.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        analysis_title.alignment = Alignment(horizontal="center")
        
        # Calculate trend
        first_week_qty = sum(daily_sales[d]["quantity"] for d in sorted_dates[:min(7, len(sorted_dates))])
        last_week_qty = sum(daily_sales[d]["quantity"] for d in sorted_dates[-min(7, len(sorted_dates)):])
        
        if last_week_qty > first_week_qty:
            trend = "📈 بڑھتا ہوا رجحان (Increasing Trend) - کاروبار ترقی کر رہا ہے"
            trend_color = "10b981"
        elif last_week_qty < first_week_qty:
            trend = "📉 گرتا ہوا رجحان (Decreasing Trend) - کاروبار میں کمی آرہی ہے"
            trend_color = "ef4444"
        else:
            trend = "➡️ مستحکم رجحان (Stable Trend) - کاروبار مستحکم ہے"
            trend_color = "f59e0b"
        
        trend_row = analysis_row + 1
        ws5.merge_cells(f'A{trend_row}:E{trend_row}')
        trend_cell = ws5.cell(row=trend_row, column=1)
        trend_cell.value = f"مجموعی رجحان: {trend}"
        trend_cell.font = Font(bold=True, size=11, color=trend_color)
        trend_cell.alignment = Alignment(horizontal="center")
        
        # Best and worst days
        best_day = max(daily_sales.items(), key=lambda x: x[1]["revenue"])
        worst_day = min(daily_sales.items(), key=lambda x: x[1]["revenue"])
        
        best_row = trend_row + 1
        ws5.cell(row=best_row, column=1, value="بہترین دن (Best Day):")
        ws5.cell(row=best_row, column=2, value=best_day[0])
        ws5.cell(row=best_row, column=3, value=f"Rs. {best_day[1]['revenue']:,.2f}")
        ws5.cell(row=best_row, column=4, value=f"{best_day[1]['quantity']} یونٹس")
        
        worst_row = best_row + 1
        ws5.cell(row=worst_row, column=1, value="کمزور ترین دن (Worst Day):")
        ws5.cell(row=worst_row, column=2, value=worst_day[0])
        ws5.cell(row=worst_row, column=3, value=f"Rs. {worst_day[1]['revenue']:,.2f}")
        ws5.cell(row=worst_row, column=4, value=f"{worst_day[1]['quantity']} یونٹس")
        
        for row in [best_row, worst_row]:
            for col in range(1, 5):
                cell = ws5.cell(row=row, column=col)
                cell.border = border
                if col == 1:
                    cell.font = Font(bold=True)
                    cell.fill = subheader_fill
                    cell.alignment = cell_alignment_left
                else:
                    cell.alignment = cell_alignment_right
    
    for col in ws5.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 3, 25)
        ws5.column_dimensions[column_letter].width = adjusted_width
    
    # =============================================================
    # SHEET 6: کسٹمر وار خلاصہ (Customer Summary)
    # =============================================================
    ws6 = wb.create_sheet("کسٹمر وار خلاصہ")
    
    ws6.merge_cells('A1:E1')
    header_cell6 = ws6.cell(row=1, column=1)
    header_cell6.value = "👥 کسٹمر وار فروخت کا خلاصہ"
    header_cell6.font = Font(bold=True, size=16, color="1a5a7a")
    header_cell6.alignment = Alignment(horizontal="center")
    
    ws6.merge_cells('A2:E2')
    subheader_cell6 = ws6.cell(row=2, column=1)
    subheader_cell6.value = "ہر کسٹمر کی کل خریداری کا تفصیلی جائزہ"
    subheader_cell6.font = Font(size=10, italic=True, color="666666")
    subheader_cell6.alignment = Alignment(horizontal="center")
    
    # Group by customer
    customer_sales = defaultdict(lambda: {"quantity": 0, "revenue": 0, "count": 0})
    for s in sales_list:
        customer_sales[s["customer_name"]]["quantity"] += s["quantity_sold"]
        customer_sales[s["customer_name"]]["revenue"] += s["total_amount"]
        customer_sales[s["customer_name"]]["count"] += 1
    
    customer_headers = ["درجہ", "کسٹمر کا نام", "کل خریداری مقدار", "کل رقم (Rs.)", "خریداری کی تعداد"]
    
    for col, header in enumerate(customer_headers, 1):
        cell = ws6.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    sorted_customers = sorted(customer_sales.items(), key=lambda x: x[1]["revenue"], reverse=True)
    for rank, (customer_name, data) in enumerate(sorted_customers, 1):
        row_num = 4 + rank
        
        ws6.cell(row=row_num, column=1, value=rank)
        ws6.cell(row=row_num, column=2, value=customer_name)
        ws6.cell(row=row_num, column=3, value=int(data["quantity"]))
        ws6.cell(row=row_num, column=4, value=round(data["revenue"], 2))
        ws6.cell(row=row_num, column=5, value=data["count"])
        
        for col in range(1, 6):
            cell = ws6.cell(row=row_num, column=col)
            cell.border = border
            if col in [3, 4]:
                cell.alignment = cell_alignment_right
            elif col == 2:
                cell.alignment = cell_alignment_left
            else:
                cell.alignment = cell_alignment_center
    
    if sorted_customers:
        total_row6 = 4 + len(sorted_customers) + 1
        ws6.merge_cells(f'A{total_row6}:A{total_row6}')
        total_label6 = ws6.cell(row=total_row6, column=1)
        total_label6.value = "کل"
        total_label6.font = Font(bold=True)
        total_label6.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        total_label6.alignment = Alignment(horizontal="center")
        
        ws6.cell(row=total_row6, column=3, value=int(total_qty))
        ws6.cell(row=total_row6, column=4, value=round(total_rev, 2))
        ws6.cell(row=total_row6, column=5, value=total_trans)
        
        for col in [3, 4, 5]:
            cell = ws6.cell(row=total_row6, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
            cell.border = border
            cell.alignment = cell_alignment_right
    
    for col in ws6.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 3, 25)
        ws6.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(excel_path)
    
    # Calculate top 5 items for KPI response
    top_5_list = [(name, data["quantity"], data["revenue"]) for name, data in top_5]
    
    return {
        "error": False,
        "folder": folder,
        "excel": excel_path,
        "kpi": {
            "total_quantity": total_qty,
            "total_revenue": total_rev,
            "total_transactions": total_trans,
            "unique_items": total_items,
            "avg_per_transaction": avg_per_transaction,
            "top_5_items": top_5_list,
        },
    }