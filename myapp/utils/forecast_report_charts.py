"""
Forecast Report Generator — Excel Only
====================================
Generates comprehensive Excel forecast report with:
- Item trend analysis (increasing/decreasing/stable)
- Sales forecasting based on historical data
- Only analyzes items with 3+ sales records
- Professional formatting
- Urdu/English support
"""

import os
from collections import defaultdict
from datetime import datetime
from typing import List, Dict, Tuple
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')


def _urdu_msg(text):
    """Convert message to Urdu display format for console"""
    return text


def fmt(value) -> str:
    """Format numbers nicely"""
    if value is None:
        return "0"
    try:
        n = float(value)
        return f"{int(n):,}" if n == int(n) else f"{n:,.2f}".rstrip("0").rstrip(".")
    except Exception:
        return str(value)


class SalesForecastEngine:
    def __init__(self, sales_data: List[Dict], forecast_days: int):
        self.sales_data = sales_data
        self.forecast_days = forecast_days
        self.forecast_results = {
            'increasing': [],
            'decreasing': [],
            'stable': []
        }
        self.threshold = 5  # 5% threshold for significant change
    
    def validate_data(self) -> Tuple[bool, str]:
        """Validate if data is sufficient for forecasting"""
        # Count items with sufficient sales (3+ records)
        item_sales_count = defaultdict(int)
        for sale in self.sales_data:
            item_sales_count[sale['item_name']] += 1
        
        valid_items = {item for item, count in item_sales_count.items() if count >= 3}
        
        if len(valid_items) < 2:
            return False, f"کم از کم 2 مختلف اشیاء کی ضرورت ہے جن کی کم از کم 3 سیلز ہوں۔ موجودہ: {len(valid_items)} اشیاء جن کی 3+ سیلز ہیں"
        
        total_sales = len(self.sales_data)
        if total_sales < 5:
            return False, f"کم از کم 5 سیلز ریکارڈ کی ضرورت ہے۔ موجودہ: {total_sales}"
        
        return True, "OK"
    
    def prepare_item_data(self, item_sales: List[Dict]) -> pd.DataFrame:
        """Prepare DataFrame for forecasting"""
        df = pd.DataFrame([{
            'ds': s['sale_date'],
            'y': s['quantity_sold']
        } for s in item_sales])
        df['ds'] = pd.to_datetime(df['ds'])
        df = df.groupby('ds').agg({'y': 'sum'}).reset_index()
        df = df.sort_values('ds')
        return df
    
    def calculate_trend_description(self, change_pct: float, current_avg: float, forecast_avg: float) -> str:
        """Generate detailed trend description in Urdu"""
        if change_pct > 15:
            return f"تیزی سے بڑھ رہی ہے - {change_pct:.1f}% اضافہ متوقع (موجودہ: {current_avg:.1f} → متوقع: {forecast_avg:.1f})"
        elif change_pct > 5:
            return f"آہستہ بڑھ رہی ہے - {change_pct:.1f}% اضافہ متوقع (موجودہ: {current_avg:.1f} → متوقع: {forecast_avg:.1f})"
        elif change_pct < -15:
            return f"تیزی سے گھٹ رہی ہے - {abs(change_pct):.1f}% کمی متوقع (موجودہ: {current_avg:.1f} → متوقع: {forecast_avg:.1f})"
        elif change_pct < -5:
            return f"آہستہ گھٹ رہی ہے - {abs(change_pct):.1f}% کمی متوقع (موجودہ: {current_avg:.1f} → متوقع: {forecast_avg:.1f})"
        else:
            return f"مستحکم ہے - صرف {abs(change_pct):.1f}% تبدیلی متوقع (موجودہ: {current_avg:.1f} → متوقع: {forecast_avg:.1f})"
    
    def has_sufficient_data(self, item_sales: List[Dict]) -> bool:
        """Check if item has at least 3 sales records"""
        return len(item_sales) >= 3
    
    def forecast_item(self, item_name: str, item_sales: List[Dict]) -> Dict:
        """Generate forecast for a single item (only if 3+ sales)"""
        try:
            # Skip items with less than 3 sales
            if not self.has_sufficient_data(item_sales):
                return None
            
            df = self.prepare_item_data(item_sales)
            if len(df) < 2:
                return None
            
            # Calculate current average (last 7 days or all if less)
            recent_data = df.tail(min(7, len(df)))
            current_avg = recent_data['y'].mean() if len(recent_data) > 0 else df['y'].mean()
            
            if current_avg == 0:
                current_avg = sum(s['quantity_sold'] for s in item_sales) / len(item_sales)
                if current_avg == 0:
                    return None
            
            # Calculate trend with linear regression
            if len(df) >= 3:
                x = np.arange(len(df))
                y = df['y'].values
                slope = np.polyfit(x, y, 1)[0]
                trend_per_day = slope
                forecast_avg = current_avg + (trend_per_day * self.forecast_days)
                forecast_avg = max(0, forecast_avg)
            else:
                forecast_avg = current_avg
            
            # Determine trend based on threshold
            change_pct = ((forecast_avg - current_avg) / current_avg) * 100 if current_avg > 0 else 0
            
            if change_pct > self.threshold:
                trend_type = 'increasing'
            elif change_pct < -self.threshold:
                trend_type = 'decreasing'
            else:
                trend_type = 'stable'
            
            # Calculate total sales and other metrics
            total_quantity = sum(s['quantity_sold'] for s in item_sales)
            total_revenue = sum(s['total_amount'] for s in item_sales)
            avg_price = total_revenue / total_quantity if total_quantity > 0 else 0
            
            return {
                'item_name': item_name,
                'trend': trend_type,
                'change_percentage': abs(change_pct),
                'current_avg': current_avg,
                'forecast_avg': forecast_avg,
                'total_sales': total_quantity,
                'total_revenue': total_revenue,
                'avg_price': avg_price,
                'unit': item_sales[0].get('item_unit', 'عدد'),
                'sales_count': len(item_sales),
                'trend_description': self.calculate_trend_description(change_pct, current_avg, forecast_avg)
            }
            
        except Exception as e:
            print(f"Error forecasting {item_name}: {e}")
            return None
    
    def run_forecast(self) -> Dict:
        """Run forecast for all items (only items with 3+ sales)"""
        item_sales_map = defaultdict(list)
        for sale in self.sales_data:
            item_sales_map[sale['item_name']].append(sale)
        
        skipped_items = []
        for item_name, item_sales in item_sales_map.items():
            if self.has_sufficient_data(item_sales):
                result = self.forecast_item(item_name, item_sales)
                if result:
                    self.forecast_results[result['trend']].append(result)
            else:
                skipped_items.append(f"{item_name} ({len(item_sales)} سیلز)")
        
        if skipped_items:
            print(f"⚠️ مندرجہ ذیل اشیاء کو چھوڑ دیا گیا (3 سے کم سیلز): {', '.join(skipped_items[:5])}")
        
        # Sort by change percentage
        self.forecast_results['increasing'].sort(key=lambda x: x['change_percentage'], reverse=True)
        self.forecast_results['decreasing'].sort(key=lambda x: x['change_percentage'], reverse=True)
        self.forecast_results['stable'].sort(key=lambda x: x['change_percentage'], reverse=True)
        
        return self.forecast_results


class ForecastExcelGenerator:
    def __init__(self, output_path: str):
        self.output_path = output_path
    
    def generate_excel(self, forecasts: Dict, forecast_days: int, total_items: int, 
                       total_quantity: float, total_revenue: float, skipped_items: List[str] = None) -> str:
        """Generate comprehensive Excel forecast report"""
        
        wb = Workbook()
        
        # Style definitions
        header_font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
        header_fill = PatternFill(start_color="2C7DA0", end_color="2C7DA0", fill_type="solid")
        subheader_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        cell_alignment_center = Alignment(horizontal="center", vertical="center")
        cell_alignment_right = Alignment(horizontal="right", vertical="center")
        cell_alignment_left = Alignment(horizontal="left", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # =============================================================
        # SHEET 1: پیشن گوئی کا خلاصہ (Forecast Summary)
        # =============================================================
        ws1 = wb.active
        ws1.title = "پیشن گوئی کا خلاصہ"
        
        # Header
        ws1.merge_cells('A1:H1')
        header_cell = ws1.cell(row=1, column=1)
        header_cell.value = f"📊 فروخت کی پیشن گوئی رپورٹ - {datetime.now().strftime('%Y-%m-%d')}"
        header_cell.font = Font(bold=True, size=16, color="1a5a7a", name="Arial")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws1.merge_cells('A2:H2')
        subheader_cell = ws1.cell(row=2, column=1)
        subheader_cell.value = f"اگلے {forecast_days} دنوں کے لیے فروخت کی پیشن گوئی (صرف ان اشیاء کا تجزیہ جن کی 3+ سیلز ہیں)"
        subheader_cell.font = Font(size=10, italic=True, color="666666", name="Arial")
        subheader_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Summary statistics
        inc_count = len(forecasts.get('increasing', []))
        dec_count = len(forecasts.get('decreasing', []))
        stable_count = len(forecasts.get('stable', []))
        analyzed_items = inc_count + dec_count + stable_count
        
        summary_data = [
            ["تفصیل (Description)", "تعداد (Count)", "فیصد (Percentage)", "نوٹس (Notes)"],
            ["📈 بڑھنے والی اشیاء (Increasing Items)", inc_count, f"{(inc_count/analyzed_items*100):.1f}%" if analyzed_items > 0 else "0%", "ان اشیاء پر اسٹاک بڑھائیں"],
            ["📉 گھٹنے والی اشیاء (Decreasing Items)", dec_count, f"{(dec_count/analyzed_items*100):.1f}%" if analyzed_items > 0 else "0%", "ان اشیاء پر پروموشن دیں"],
            ["➡️ مستحکم اشیاء (Stable Items)", stable_count, f"{(stable_count/analyzed_items*100):.1f}%" if analyzed_items > 0 else "0%", "موجودہ حکمت عملی جاری رکھیں"],
            ["✅ کل تجزیہ شدہ اشیاء (Total Analyzed Items)", analyzed_items, "100%", "جن کی 3+ سیلز ہیں"],
        ]
        
        row = 4
        for data_row in summary_data:
            for col_idx, value in enumerate(data_row, 1):
                cell = ws1.cell(row=row, column=col_idx)
                cell.value = value
                cell.border = border
                
                if row == 4:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                else:
                    if col_idx == 1:
                        cell.font = Font(bold=True)
                        cell.fill = subheader_fill
                        cell.alignment = cell_alignment_left
                    elif col_idx in [2, 3]:
                        cell.alignment = cell_alignment_right
                    else:
                        cell.alignment = cell_alignment_left
            row += 1
        
        # Add note about skipped items
        if skipped_items:
            row += 1
            ws1.merge_cells(f'A{row}:H{row}')
            note_cell = ws1.cell(row=row, column=1)
            note_cell.value = f"⚠️ نوٹ: {len(skipped_items)} اشیاء کو تجزیہ سے خارج کر دیا گیا کیونکہ ان کی 3 سے کم سیلز ہیں۔"
            note_cell.font = Font(color="ef4444", size=9)
            note_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            note_cell.alignment = Alignment(horizontal="center")
        
        # Overall forecast recommendation
        row += 2
        ws1.merge_cells(f'A{row}:H{row}')
        rec_cell = ws1.cell(row=row, column=1)
        
        if analyzed_items == 0:
            recommendation = f"⚠️ کوئی بھی آئٹم تجزیہ کے لیے موزوں نہیں ہے۔ براہ کرم مزید سیلز ریکارڈز شامل کریں (کم از کم 3 سیلز فی آئٹم)"
        elif inc_count > dec_count:
            recommendation = f"✅ مجموعی رجحان: بہتر ہے! {inc_count} اشیاء کی فروخت بڑھ رہی ہے۔ ان اشیاء پر اسٹاک بڑھانے کی تجویز ہے۔"
        elif dec_count > inc_count:
            recommendation = f"⚠️ مجموعی رجحان: تشویشناک ہے! {dec_count} اشیاء کی فروخت گھٹ رہی ہے۔ ان اشیاء پر پروموشن اور ڈسکاؤنٹ دیں۔"
        else:
            recommendation = f"➡️ مجموعی رجحان: مستحکم ہے۔ موجودہ حکمت عملی کو برقرار رکھیں۔"
        
        rec_cell.value = recommendation
        rec_cell.font = Font(bold=True, size=11, color="2C7DA0")
        rec_cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        rec_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust columns
        for col in ws1.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 40)
            ws1.column_dimensions[column_letter].width = adjusted_width
        
        # =============================================================
        # SHEET 2: بڑھنے والی اشیاء (Increasing Items)
        # =============================================================
        if forecasts.get('increasing'):
            ws2 = wb.create_sheet("بڑھنے والی اشیاء")
            
            ws2.merge_cells('A1:H1')
            header_cell2 = ws2.cell(row=1, column=1)
            header_cell2.value = f"📈 بڑھنے والی اشیاء (ٹاپ {min(10, len(forecasts['increasing']))})"
            header_cell2.font = Font(bold=True, size=16, color="10b981", name="Arial")
            header_cell2.alignment = Alignment(horizontal="center", vertical="center")
            
            ws2.merge_cells('A2:H2')
            subheader_cell2 = ws2.cell(row=2, column=1)
            subheader_cell2.value = f"یہ اشیاء اگلے {forecast_days} دنوں میں فروخت میں اضافہ دکھائیں گی۔ ان پر اسٹاک بڑھانے کی تجویز ہے۔"
            subheader_cell2.font = Font(size=10, italic=True, color="666666", name="Arial")
            subheader_cell2.alignment = Alignment(horizontal="center", vertical="center")
            
            inc_headers = [
                "درجہ", "آئٹم کا نام", "یونٹ", "موجودہ اوسط", "متوقع اوسط",
                "متوقع اضافہ (%)", "تفصیلی تجزیہ", "تجاویز"
            ]
            
            for col, header in enumerate(inc_headers, 1):
                cell = ws2.cell(row=4, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
                cell.alignment = header_alignment
                cell.border = border
            
            for idx, item in enumerate(forecasts['increasing'][:10], 1):
                row_num = 4 + idx
                
                if item['change_percentage'] > 15:
                    recommendation = f"⚠️ بہت تیزی سے بڑھ رہی ہے! فوری طور پر اسٹاک ڈبل کریں"
                elif item['change_percentage'] > 10:
                    recommendation = f"👍 اچھی بڑھوتری! اسٹاک 50% بڑھائیں"
                elif item['change_percentage'] > 5:
                    recommendation = f"✅ معمولی اضافہ! اسٹاک 25% بڑھائیں"
                else:
                    recommendation = f"➡️ معمولی تبدیلی، موجودہ اسٹاک برقرار رکھیں"
                
                ws2.cell(row=row_num, column=1, value=idx)
                ws2.cell(row=row_num, column=2, value=item['item_name'])
                ws2.cell(row=row_num, column=3, value=item['unit'])
                ws2.cell(row=row_num, column=4, value=round(item['current_avg'], 1))
                ws2.cell(row=row_num, column=5, value=round(item['forecast_avg'], 1))
                ws2.cell(row=row_num, column=6, value=f"+{item['change_percentage']:.1f}%")
                ws2.cell(row=row_num, column=7, value=item['trend_description'])
                ws2.cell(row=row_num, column=8, value=recommendation)
                
                for col in range(1, 9):
                    cell = ws2.cell(row=row_num, column=col)
                    cell.border = border
                    if col in [4, 5, 6]:
                        cell.alignment = cell_alignment_right
                    elif col == 2:
                        cell.alignment = cell_alignment_left
                        cell.font = Font(bold=True)
                    else:
                        cell.alignment = cell_alignment_left
                    
                    if col == 6:
                        cell.font = Font(bold=True, color="10b981")
            
            for col in ws2.columns:
                max_length = 0
                column_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 3, 45)
                ws2.column_dimensions[column_letter].width = adjusted_width
        
        # =============================================================
        # SHEET 3: گھٹنے والی اشیاء (Decreasing Items)
        # =============================================================
        if forecasts.get('decreasing'):
            ws3 = wb.create_sheet("گھٹنے والی اشیاء")
            
            ws3.merge_cells('A1:H1')
            header_cell3 = ws3.cell(row=1, column=1)
            header_cell3.value = f"📉 گھٹنے والی اشیاء (ٹاپ {min(10, len(forecasts['decreasing']))})"
            header_cell3.font = Font(bold=True, size=16, color="ef4444", name="Arial")
            header_cell3.alignment = Alignment(horizontal="center", vertical="center")
            
            ws3.merge_cells('A2:H2')
            subheader_cell3 = ws3.cell(row=2, column=1)
            subheader_cell3.value = f"یہ اشیاء اگلے {forecast_days} دنوں میں فروخت میں کمی دکھائیں گی۔ ان پر پروموشن اور ڈسکاؤنٹ دیں۔"
            subheader_cell3.font = Font(size=10, italic=True, color="666666", name="Arial")
            subheader_cell3.alignment = Alignment(horizontal="center", vertical="center")
            
            dec_headers = [
                "درجہ", "آئٹم کا نام", "یونٹ", "موجودہ اوسط", "متوقع اوسط",
                "متوقع کمی (%)", "تفصیلی تجزیہ", "تجاویز"
            ]
            
            for col, header in enumerate(dec_headers, 1):
                cell = ws3.cell(row=4, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid")
                cell.alignment = header_alignment
                cell.border = border
            
            for idx, item in enumerate(forecasts['decreasing'][:10], 1):
                row_num = 4 + idx
                
                if item['change_percentage'] > 15:
                    recommendation = f"⚠️ بہت تیزی سے گر رہی ہے! فوری پروموشن اور 30% ڈسکاؤنٹ دیں"
                elif item['change_percentage'] > 10:
                    recommendation = f"👍 تشویشناک کمی! پروموشن اور 20% ڈسکاؤنٹ دیں"
                elif item['change_percentage'] > 5:
                    recommendation = f"✅ معمولی کمی! 10% ڈسکاؤنٹ یا BOGO آفر دیں"
                else:
                    recommendation = f"➡️ معمولی تبدیلی، وجہ معلوم کریں"
                
                ws3.cell(row=row_num, column=1, value=idx)
                ws3.cell(row=row_num, column=2, value=item['item_name'])
                ws3.cell(row=row_num, column=3, value=item['unit'])
                ws3.cell(row=row_num, column=4, value=round(item['current_avg'], 1))
                ws3.cell(row=row_num, column=5, value=round(item['forecast_avg'], 1))
                ws3.cell(row=row_num, column=6, value=f"-{item['change_percentage']:.1f}%")
                ws3.cell(row=row_num, column=7, value=item['trend_description'])
                ws3.cell(row=row_num, column=8, value=recommendation)
                
                for col in range(1, 9):
                    cell = ws3.cell(row=row_num, column=col)
                    cell.border = border
                    if col in [4, 5, 6]:
                        cell.alignment = cell_alignment_right
                    elif col == 2:
                        cell.alignment = cell_alignment_left
                        cell.font = Font(bold=True)
                    else:
                        cell.alignment = cell_alignment_left
                    
                    if col == 6:
                        cell.font = Font(bold=True, color="ef4444")
            
            for col in ws3.columns:
                max_length = 0
                column_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 3, 45)
                ws3.column_dimensions[column_letter].width = adjusted_width
        
        # =============================================================
        # SHEET 4: تمام اشیاء کا مکمل تجزیہ (Complete Analysis)
        # =============================================================
        ws4 = wb.create_sheet("تمام اشیاء کا مکمل تجزیہ")
        
        ws4.merge_cells('A1:I1')
        header_cell4 = ws4.cell(row=1, column=1)
        header_cell4.value = "📋 تمام اشیاء کا مکمل پیشن گوئی تجزیہ"
        header_cell4.font = Font(bold=True, size=16, color="1a5a7a", name="Arial")
        header_cell4.alignment = Alignment(horizontal="center", vertical="center")
        
        ws4.merge_cells('A2:I2')
        subheader_cell4 = ws4.cell(row=2, column=1)
        subheader_cell4.value = "ہر آئٹم کے لیے تفصیلی پیشن گوئی اور تجاویز (صرف وہ اشیاء جن کی 3+ سیلز ہیں)"
        subheader_cell4.font = Font(size=10, italic=True, color="666666", name="Arial")
        subheader_cell4.alignment = Alignment(horizontal="center", vertical="center")
        
        all_headers = [
            "درجہ", "آئٹم کا نام", "رجحان", "موجودہ اوسط", "متوقع اوسط",
            "تبدیلی (%)", "کل فروخت", "تفصیلی تجزیہ", "تجاویز"
        ]
        
        for col, header in enumerate(all_headers, 1):
            cell = ws4.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        all_items = []
        for item in forecasts.get('increasing', []):
            all_items.append({**item, 'trend_urdu': '📈 بڑھنے والی', 'trend_color': '10b981'})
        for item in forecasts.get('decreasing', []):
            all_items.append({**item, 'trend_urdu': '📉 گھٹنے والی', 'trend_color': 'ef4444'})
        for item in forecasts.get('stable', []):
            all_items.append({**item, 'trend_urdu': '➡️ مستحکم', 'trend_color': 'f59e0b'})
        
        all_items.sort(key=lambda x: x['change_percentage'], reverse=True)
        
        for idx, item in enumerate(all_items, 1):
            row_num = 4 + idx
            
            if item['trend'] == 'increasing':
                if item['change_percentage'] > 15:
                    recommendation = "⚠️ فوری اسٹاک ڈبل کریں"
                elif item['change_percentage'] > 10:
                    recommendation = "👍 اسٹاک 50% بڑھائیں"
                elif item['change_percentage'] > 5:
                    recommendation = "✅ اسٹاک 25% بڑھائیں"
                else:
                    recommendation = "موجودہ اسٹاک برقرار رکھیں"
            elif item['trend'] == 'decreasing':
                if item['change_percentage'] > 15:
                    recommendation = "⚠️ فوری پروموشن + 30% ڈسکاؤنٹ"
                elif item['change_percentage'] > 10:
                    recommendation = "👍 پروموشن + 20% ڈسکاؤنٹ"
                elif item['change_percentage'] > 5:
                    recommendation = "✅ 10% ڈسکاؤنٹ یا BOGO"
                else:
                    recommendation = "وجہ معلوم کریں"
            else:
                recommendation = "موجودہ حکمت عملی برقرار رکھیں"
            
            ws4.cell(row=row_num, column=1, value=idx)
            ws4.cell(row=row_num, column=2, value=item['item_name'])
            ws4.cell(row=row_num, column=3, value=item['trend_urdu'])
            ws4.cell(row=row_num, column=4, value=round(item['current_avg'], 1))
            ws4.cell(row=row_num, column=5, value=round(item['forecast_avg'], 1))
            
            change_sign = "+" if item['trend'] == 'increasing' else "-"
            ws4.cell(row=row_num, column=6, value=f"{change_sign}{item['change_percentage']:.1f}%")
            ws4.cell(row=row_num, column=7, value=int(item['total_sales']))
            ws4.cell(row=row_num, column=8, value=item['trend_description'])
            ws4.cell(row=row_num, column=9, value=recommendation)
            
            for col in range(1, 10):
                cell = ws4.cell(row=row_num, column=col)
                cell.border = border
                if col in [4, 5, 6, 7]:
                    cell.alignment = cell_alignment_right
                elif col == 2:
                    cell.alignment = cell_alignment_left
                    cell.font = Font(bold=True)
                else:
                    cell.alignment = cell_alignment_left
                
                if col == 3:
                    cell.font = Font(bold=True, color=item['trend_color'])
                elif col == 6:
                    if item['trend'] == 'increasing':
                        cell.font = Font(bold=True, color="10b981")
                    elif item['trend'] == 'decreasing':
                        cell.font = Font(bold=True, color="ef4444")
        
        for col in ws4.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 45)
            ws4.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(self.output_path)
        return self.output_path


async def generate_forecast_report(report_id: str, sales_data: List[Dict], forecast_days: int) -> Dict:
    """Main function to generate complete forecast report (Excel only)"""
    
    try:
        print(f"📊 پیشن گوئی رپورٹ تیار کر رہا ہے: {len(sales_data)} سیلز ریکارڈ, {forecast_days} دن")
        
        # Create output folder
        output_folder = f"forecast_reports/report_{report_id}"
        os.makedirs(output_folder, exist_ok=True)
        
        # Run forecast
        forecaster = SalesForecastEngine(sales_data, forecast_days)
        
        # Validate data
        is_valid, message = forecaster.validate_data()
        if not is_valid:
            print(f"❌ {message}")
            return {"error": True, "message": message}
        
        # Generate forecasts
        forecasts = forecaster.run_forecast()
        
        inc_count = len(forecasts.get('increasing', []))
        dec_count = len(forecasts.get('decreasing', []))
        stable_count = len(forecasts.get('stable', []))
        total_items = inc_count + dec_count + stable_count
        
        # Find skipped items
        item_sales_count = defaultdict(int)
        for sale in sales_data:
            item_sales_count[sale['item_name']] += 1
        skipped_items = [f"{item} ({count} سیلز)" for item, count in item_sales_count.items() if count < 3]
        
        # Calculate total quantities
        total_quantity = sum(s['quantity_sold'] for s in sales_data)
        total_revenue = sum(s['total_amount'] for s in sales_data)
        
        print(f"📈 پیشن گوئی کے نتائج: بڑھنے والی: {inc_count}, گھٹنے والی: {dec_count}, مستحکم: {stable_count}")
        if skipped_items:
            print(f"⚠️ چھوڑی گئی اشیاء (3 سے کم سیلز): {len(skipped_items)}")
        
        # Generate Excel only
        excel_path = os.path.join(output_folder, f"forecast_report_{forecast_days}days.xlsx")
        excel_gen = ForecastExcelGenerator(excel_path)
        excel_gen.generate_excel(forecasts, forecast_days, total_items, total_quantity, total_revenue, skipped_items)
        print(f"✅ ایکسل فائل محفوظ: {excel_path}")
        
        print("🎉 تمام فائلیں کامیابی سے تیار ہو گئیں")
        
        return {
            "error": False,
            "report_id": report_id,
            "output_folder": output_folder,
            "excel_path": excel_path,
            "forecast_summary": {
                "total_items_analyzed": total_items,
                "increasing_count": inc_count,
                "decreasing_count": dec_count,
                "stable_count": stable_count,
                "forecast_days": forecast_days,
                "total_quantity": total_quantity,
                "total_revenue": total_revenue,
                "skipped_items_count": len(skipped_items)
            }
        }
        
    except Exception as e:
        print(f"❌ پیشن گوئی رپورٹ میں خرابی: {e}")
        import traceback
        traceback.print_exc()
        return {"error": True, "message": str(e)}


async def delete_forecast_report(folder_path: str) -> Dict:
    """Delete forecast report folder with Windows compatibility"""
    import shutil
    import stat
    import time
    
    if not os.path.exists(folder_path):
        return {"status": "not_found", "path": folder_path}
    
    def _remove_readonly(func, path, excinfo):
        """Error handler for shutil.rmtree on Windows."""
        try:
            os.chmod(path, stat.S_IWRITE)
            func(path)
        except Exception:
            pass
    
    for attempt in range(3):
        try:
            shutil.rmtree(folder_path, ignore_errors=False, onerror=_remove_readonly)
            print(f"✅ Deleted folder: {folder_path}")
            return {"status": "deleted", "path": folder_path}
        except PermissionError:
            if attempt == 2:
                print(f"❌ Failed to delete {folder_path} after 3 attempts")
                return {"status": "failed", "path": folder_path, "error": "Permission denied"}
            time.sleep(0.3 * (attempt + 1))
        except Exception as e:
            print(f"❌ Error deleting {folder_path}: {e}")
            return {"status": "failed", "path": folder_path, "error": str(e)}
    
    return {"status": "failed", "path": folder_path, "error": "Unknown error"}